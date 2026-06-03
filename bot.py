import os
from datetime import datetime
from openpyxl import load_workbook
from telegram import Bot
from telegram.error import TelegramError
import asyncio
import schedule
import time

# ===== НАСТРОЙКИ (ЗАМЕНИТЕ НА СВОИ!) =====
TOKEN = "8738210675:AAEcUNmlqYCatgimMypA8XnTR7d3dMxd9qY"           # Вставьте токен от BotFather
CHAT_ID = -1003910935617                 # Вставьте ваш ID (число, без кавычек!)
EXCEL_FILE = "graphic.xlsx"        # Имя файла
# ===========================================

# Строим полный путь к файлу (там же, где скрипт)
script_dir = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(script_dir, EXCEL_FILE)

def get_today_employees():
    """Читает Excel и возвращает ВСЕХ сотрудников на сегодня"""
    today = datetime.now().strftime("%d.%m")
    
    if not os.path.exists(EXCEL_FILE):
        return f"❌ Ошибка: файл {EXCEL_FILE} не найден!"
    
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
    except Exception as e:
        return f"❌ Ошибка при чтении файла: {e}"
    
    # Список для хранения всех сотрудников на сегодня
    employees_today = []
    
    for row in range(2, sheet.max_row + 1):
        date_cell = sheet.cell(row=row, column=1).value
        employee_cell = sheet.cell(row=row, column=2).value
        time_cell = sheet.cell(row=row, column=3).value
        
        # Преобразуем дату из Excel в строку
        if hasattr(date_cell, 'strftime'):
            date_cell = date_cell.strftime("%d.%m")
        
        # Если дата совпадает с сегодняшней — добавляем в список
        if str(date_cell) == today:
            # Формируем время начала работы
            if time_cell:
                if hasattr(time_cell, 'strftime'):
                    start_time = time_cell.strftime("%H:%M")
                else:
                    start_time = str(time_cell)
            else:
                start_time = "не указано"
            
            employees_today.append(f"👤 {employee_cell} — {start_time}")
    
    # Формируем итоговое сообщение
    if employees_today:
        # Сортируем по времени (опционально, для удобства)
        # employees_today.sort()  # раскомментировать если нужно сортировать
        
        all_employees = "\n".join(employees_today)
        return (f"📅 Сегодня {today}\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"{all_employees}")
    else:
        return (f"📅 Сегодня {today}\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"⚠️ Сотрудники на эту дату не найдены!\n"
                f"Проверьте, что дата '{today}' есть в таблице.")

async def send_daily_message():
    """Отправляет сообщение в Telegram"""
    bot = Bot(token=TOKEN)
    message = get_today_employees()
    
    try:
        await bot.send_message(chat_id=CHAT_ID, text=message)
        print(f"[✓] Сообщение отправлено: {message}")
    except TelegramError as e:
        print(f"[✗] Ошибка: {e}")

def send_sync():
    asyncio.run(send_daily_message())

if __name__ == "__main__":
    print("🚀 Запуск бота...")
    print(f"📁 Файл таблицы: {EXCEL_FILE}")
    print(f"📂 Существует: {'Да' if os.path.exists(EXCEL_FILE) else 'Нет'}")
    
    # Отправляем сразу для проверки
    print("\n📨 Отправка сообщения...")
    asyncio.run(send_daily_message())
    
    print("\n✅ Бот запущен и будет отправлять уведомления каждый день в 5:45")
    print("🛑 Для остановки нажмите Ctrl+C\n")
    
    schedule.every().day.at("05:45").do(send_sync)
    
    while True:
        schedule.run_pending()
        time.sleep(60)
